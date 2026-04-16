"""
Excel output module — writes tractian_leads.xlsx with 3 professional sheets.
"""
from __future__ import annotations
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .config import OUTPUTS_DIR
from .logger import get_logger

log = get_logger("output_xlsx")

# Color scheme
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AMBER_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="0A0F1C", end_color="0A0F1C", fill_type="solid")
HEADER_FONT = Font(bold=True, color="00D4FF", size=11)
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD'),
)

COLUMNS = [
    ("Company", 20), ("Website", 18), ("ICP Score", 10),
    ("Location", 30), ("City", 18), ("State/Region", 15),
    ("Country", 15), ("Facility Type", 20), ("Classification Basis", 35),
    ("Confidence", 12), ("Needs Verification", 16), ("Source URL", 30),
    ("Source Type", 12), ("Source Count", 12), ("Date Collected", 14),
]


def _score_fill(score: int) -> PatternFill:
    if score >= 8:
        return GREEN_FILL
    elif score >= 5:
        return AMBER_FILL
    return RED_FILL


def _style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def write_xlsx(rows: list[dict], path: Path | None = None) -> Path:
    """Write final rows to XLSX with 3 formatted sheets."""
    path = path or (OUTPUTS_DIR / "tractian_leads.xlsx")
    wb = Workbook()

    # ── Sheet 1: All Leads ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "All Leads"

    # Headers
    for i, (name, width) in enumerate(COLUMNS, 1):
        ws1.cell(row=1, column=i, value=name)
        ws1.column_dimensions[get_column_letter(i)].width = width
    _style_header(ws1, len(COLUMNS))

    # Sort by ICP score desc
    sorted_rows = sorted(rows, key=lambda r: r.get("icp_score", 0), reverse=True)

    for row_idx, row in enumerate(sorted_rows, 2):
        score = row.get("icp_score", 0)
        values = [
            row.get("company_name", ""),
            row.get("website", ""),
            score,
            row.get("facility_location", ""),
            row.get("city", ""),
            row.get("state_region", ""),
            row.get("country", ""),
            row.get("facility_type", ""),
            row.get("classification_basis", "")[:100],
            row.get("confidence", ""),
            str(row.get("needs_verification", "")),
            row.get("source_url", ""),
            row.get("source_type", ""),
            row.get("source_count", 0),
            row.get("date_collected", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx == 3:
                cell.fill = _score_fill(score)
            elif row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Freeze header
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(sorted_rows)+1}"

    # ── Sheet 2: Company Summary ────────────────────────────────────
    ws2 = wb.create_sheet("Company Summary")
    summary_cols = [
        ("Company", 22), ("ICP Score", 10), ("Score Confidence", 14),
        ("Total Facilities", 14), ("Facility Types", 35),
        ("Plain English Summary", 60), ("Data Quality", 12),
    ]
    for i, (name, width) in enumerate(summary_cols, 1):
        ws2.cell(row=1, column=i, value=name)
        ws2.column_dimensions[get_column_letter(i)].width = width
    _style_header(ws2, len(summary_cols))

    company_data = {}
    for row in sorted_rows:
        cn = row.get("company_name", "")
        if cn not in company_data:
            company_data[cn] = {"rows": [], "score": row.get("icp_score", 0), "breakdown": row.get("score_breakdown")}
        company_data[cn]["rows"].append(row)

    for row_idx, (cn, data) in enumerate(
        sorted(company_data.items(), key=lambda x: x[1]["score"], reverse=True), 2
    ):
        facility_rows = data["rows"]
        types = list(set(r.get("facility_type", "Unknown") for r in facility_rows))
        breakdown = data["breakdown"]
        if isinstance(breakdown, str):
            try:
                breakdown = json.loads(breakdown)
            except Exception:
                breakdown = {}
        plain = breakdown.get("plain_english", "") if isinstance(breakdown, dict) else ""
        conf = breakdown.get("score_confidence", "N/A") if isinstance(breakdown, dict) else "N/A"
        high_conf = sum(1 for r in facility_rows if r.get("confidence") == "HIGH")
        quality = "HIGH" if high_conf > len(facility_rows) * 0.5 else ("MED" if high_conf > 0 else "LOW")

        values = [
            cn, data["score"], conf,
            len(facility_rows), ", ".join(types[:5]),
            plain[:200], quality,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx == 2:
                cell.fill = _score_fill(data["score"])
            elif row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    ws2.freeze_panes = "A2"

    # ── Sheet 3: High Value Targets ────────────────────────────────
    ws3 = wb.create_sheet("High Value Targets")
    hvt_cols = [
        ("Company", 22), ("ICP Score", 10), ("Top Facility", 30),
        ("Facility Type", 18), ("Country", 15), ("Confidence", 12),
        ("# Facilities", 12),
    ]
    for i, (name, width) in enumerate(hvt_cols, 1):
        ws3.cell(row=1, column=i, value=name)
        ws3.column_dimensions[get_column_letter(i)].width = width
    _style_header(ws3, len(hvt_cols))

    row_idx = 2
    for cn, data in sorted(company_data.items(), key=lambda x: x[1]["score"], reverse=True):
        if data["score"] < 8:
            continue
        best = max(data["rows"], key=lambda r: {"HIGH": 3, "MED": 2, "LOW": 1, "ESTIMATED": 0}.get(r.get("confidence", "LOW"), 0))
        values = [
            cn, data["score"],
            best.get("facility_location", ""),
            best.get("facility_type", ""),
            best.get("country", ""),
            best.get("confidence", ""),
            len(data["rows"]),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx == 2:
                cell.fill = GREEN_FILL
            elif row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL
        row_idx += 1

    ws3.freeze_panes = "A2"

    wb.save(path)
    log.info(f"XLSX written: {path} ({len(rows)} rows, 3 sheets)")
    return path
